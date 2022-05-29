import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(sheet['A1'])
# returning the defined cell
# <Cell 'Sheet1'.A1>

print(sheet['A1'].value)
# getting the value from the cell
# 2015-04-05 13:34:02

c = sheet['B1']
print(c.value)
# getting the value of B1 - Apples

print(f"Cell {c.coordinate} is {c.value}")
# c.coordinated returns the name of the cell --> B1
# c.values returns the value of B1 which is Apples
# printing message containing this information

print(sheet['C1'].value)
# printing the value of C1

print(sheet.cell(row=1, column=2))
# returns: <Cell 'Sheet1'.B1>
print(sheet.cell(row=1, column=2).value)
# returns value

for i in range(1, 8, 2):
    # going through every other row
    print(i, sheet.cell(row=i, column=2).value)
    # i is the odd numbers --> 1, 3, 5, 7
    # the rest is getting the value of each section of B1

print(sheet.max_row)
# printing the highest row number

print(sheet.max_column)
# printing the highest column numbers