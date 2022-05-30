import openpyxl

wb = openpyxl.Workbook()
# creates a new workbook
print(wb.sheetnames)
# name is default 'Sheet'
sheet = wb.active
# sheet is defined as the active sheet
print(sheet.title)
# printing title as string
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)
print(sheet.title)


# changing name on sheet
# import openpyxl
# wb=openpyxl.load_workbook('example.xlsx')
# sheet=wb.active
# sheet.title='Spam Spam Spam' (changing the name of the sheet to this)
# wb.save('example_copy.xlsx') (saving the file with the new sheet name as a
# new file called copy)
