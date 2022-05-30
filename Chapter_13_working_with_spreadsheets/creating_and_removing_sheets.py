import openpyxl

wb = openpyxl.Workbook()
# creating a new excel document

print(wb.sheetnames)
# default name 'Sheet'

wb.create_sheet()
# creating a new sheet
print(wb.sheetnames)
# default name 'Sheet 1'

wb.create_sheet(index=0, title='First Sheet')
# creating a new sheet at index 0 called 'First Sheet'
print(wb.sheetnames)

wb.create_sheet(index=2, title='Middle Sheet')
# creating a new sheet at index 2 called 'Middle Sheet'
print(wb.sheetnames)

del wb['Middle Sheet']
del wb['Sheet1']
print(wb.sheetnames)