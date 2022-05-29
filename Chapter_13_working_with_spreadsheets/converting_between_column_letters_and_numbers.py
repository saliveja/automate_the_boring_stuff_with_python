import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
# translating column one to a letter --> A

print(get_column_letter(2))
# translating column one to a letter --> B

print(get_column_letter(27))
# AA

print(get_column_letter(900))
# AHP

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

print(get_column_letter(sheet.max_column))
# C (there are three columns in total, so max is C)

print(column_index_from_string('A'))
# 1

print(column_index_from_string('AA'))
# 27